#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Collect the iceland_tourism dataset

See the README file for more information.

Author: G.J.J. van den Burg
License: This file is part of TCPD, see the top-level LICENSE file.
Copyright: 2019, The Alan Turing Institute

"""

import argparse
import hashlib
import json
import openpyxl
import os
import sys
import time

from functools import wraps
from urllib.request import urlretrieve
from urllib.error import URLError

XLSX_URL = "https://web.archive.org/web/20191121170223if_/https://www.ferdamalastofa.is/static/files/ferdamalastofa/Frettamyndir/2019/nov/visitors-to-iceland-2002-2019-oct.xlsx"

MD5_XLSX = "ec777afd95b01ca901aa00475fc284e5"
MD5_JSON = "8bbac4ca95319a865f2d58ff564f063d"

NAME_XLSX = "visitors-to-iceland-2002-2019-oct.xlsx"
NAME_JSON = "iceland_tourism.json"

MONTHS = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12,
}


class ValidationError(Exception):
    def __init__(self, filename):
        self.message = (
            "Validating the file '%s' failed. \n"
            "Please raise an issue on the GitHub page for this project \n"
            "if the error persists." % filename
        )


def check_md5sum(filename, checksum):
    with open(filename, "rb") as fp:
        data = fp.read()
    h = hashlib.md5(data).hexdigest()
    return h == checksum


def validate(checksum):
    """Decorator that validates the target file."""

    def validate_decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            target = kwargs.get("target_path", None)
            if os.path.exists(target) and check_md5sum(target, checksum):
                return
            out = func(*args, **kwargs)
            if not os.path.exists(target):
                raise FileNotFoundError("Target file expected at: %s" % target)
            if not check_md5sum(target, checksum):
                raise ValidationError(target)
            return out

        return wrapper

    return validate_decorator


@validate(MD5_XLSX)
def download_xlsx(target_path=None):
    count = 0
    while count < 5:
        count += 1
        try:
            urlretrieve(XLSX_URL, target_path)
            return
        except URLError as err:
            print(
                "Error occurred (%r) when trying to download xlsx. Retrying in 5 seconds"
                % err,
                sys.stderr,
            )
            time.sleep(5)


def format_ym(year, month):
    midx = MONTHS[month]
    return "%i-%02d" % (int(year), midx)


@validate(MD5_JSON)
def write_json(xlsx_path, target_path=None):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[2]

    rows = list(ws.rows)

    # hardcoding these row indices, not worth doing it nicely
    header = rows[2]

    column_idx = [
        i
        for i, c in enumerate(header)
        if c.data_type == "n" and c.value and 2003 <= c.value < 2020
    ]

    visitors = []

    r_offset = 4
    for c in column_idx:
        for r in range(r_offset, r_offset + 12):
            cell = ws.cell(r, c + 1)
            if cell.value is None or str(cell.value) == "":
                continue
            year = header[c].value
            month = ws.cell(r, 1).value
            datestr = format_ym(year, month)
            # eliminate some observations that were not in the original dataset
            if datestr in ["2019-08", "2019-09", "2019-10"]:
                continue
            item = {"time": datestr, "value": int(cell.value)}
            visitors.append(item)

    name = "iceland_tourism"
    longname = "Iceland Tourism"

    data = {
        "name": name,
        "longname": longname,
        "n_obs": len(visitors),
        "n_dim": 1,
        "time": {
            "format": "%Y-%m",
            "index": list(range(len(visitors))),
            "raw": [v["time"] for v in visitors],
        },
        "series": [
            {
                "label": "Visitor Number",
                "type": "int",
                "raw": [v["value"] for v in visitors],
            }
        ],
    }

    with open(target_path, "w") as fp:
        json.dump(data, fp, indent="\t")


def collect(output_dir="."):
    xlsx_path = os.path.join(output_dir, NAME_XLSX)
    json_path = os.path.join(output_dir, NAME_JSON)

    download_xlsx(target_path=xlsx_path)
    write_json(xlsx_path, target_path=json_path)


def clean(output_dir="."):
    xlsx_path = os.path.join(output_dir, NAME_XLSX)
    json_path = os.path.join(output_dir, NAME_JSON)

    if os.path.exists(xlsx_path):
        os.unlink(xlsx_path)
    if os.path.exists(json_path):
        os.unlink(json_path)


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-o", "--output-dir", help="output directory to use", default="."
    )
    parser.add_argument(
        "action",
        choices=["collect", "clean"],
        help="Action to perform",
        default="collect",
        nargs="?",
    )
    return parser.parse_args()


def main(output_dir="."):
    args = parse_args()
    if args.action == "collect":
        collect(output_dir=args.output_dir)
    elif args.action == "clean":
        clean(output_dir=args.output_dir)


if __name__ == "__main__":
    main()
